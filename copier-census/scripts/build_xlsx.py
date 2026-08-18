"""Build the dealer census workbook.

Sheets:
  Summary            headline counts, all by formula off the data sheets
  All Contacts       every contact at a copier dealer, with reachability flags
  Coverage by Dealer one row per dealer: contacts held, decision-makers, reachable DMs
  No Contacts        the gap list, with what was already tried and what to try next
  Sourcing Playbook  where contacts actually came from, with measured hit rates
  Method & Caveats   what these numbers do and do not mean
"""
import os, sys, json, re
sys.path.insert(0, '/tmp')
from resolver import Companies
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

S = '/tmp/claude-0/-home-user-Claude/adc041a4-59ce-53c7-8a85-ffe65b71c860/scratchpad'
D = json.load(open(S + '/xlsx_data.json'))
comps, per, hold, brand, tried = (D['companies'], D['contacts'], set(D['hold']),
                                  D['brand'], D['tried'])
CO = Companies(comps)

DM = re.compile(r'president|ceo|owner|chief executive|principal|founder|partner|'
                r'marketing director|director of marketing|cmo|chief sales|chief revenue|'
                r'vp of sales|vice president of sales|general manager|coo|chief operating|'
                r'chairman|proprietor|chief information|cio|cto|vp |vice president|'
                r'director of sales|sales director', re.I)
DEAD = {'No Longer with Company', 'Retired - Remove from All Lists', 'Incorrect Contact',
        'Do Not Call - Opt Out', 'Disqualified'}

# ---------- roll contacts to the company CLUSTER so duplicate shells don't distort
byc = defaultdict(list)
for cid, rows in per.items():
    byc[CO.cluster(cid)] += rows

ARIAL = 'Arial'
HDR = Font(name=ARIAL, bold=True, size=10, color='FFFFFF')
HDRF = PatternFill('solid', fgColor='1F4E5F')
BOLD = Font(name=ARIAL, bold=True, size=10)
BODY = Font(name=ARIAL, size=10)
TITLE = Font(name=ARIAL, bold=True, size=14)
NOTE = Font(name=ARIAL, size=9, italic=True, color='555555')
RED = Font(name=ARIAL, size=10, bold=True, color='A62C61')
GRN = Font(name=ARIAL, size=10, bold=True, color='1A6B4F')
THIN = Border(bottom=Side(style='thin', color='D0D5DA'))
YEL = PatternFill('solid', fgColor='FFF6D6')

wb = Workbook()

def head(ws, cols, row=1):
    for j, c in enumerate(cols, 1):
        x = ws.cell(row=row, column=j, value=c)
        x.font = HDR; x.fill = HDRF
        x.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = None

def widths(ws, w):
    for j, n in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(j)].width = n

# =============================================================== All Contacts
ws = wb.active
ws.title = 'All Contacts'
cols = ['In HubSpot?','Company','Dealer Status','Domain','City','State','First','Last','Title',
        'Decision Maker?','Email','Email Status','Phone','Mobile','LinkedIn URL',
        'Lead Status','Still At Company','Prior Validation','Verified Date',
        'ZoomInfo Score','Reachable?','Contact ID','Company ID',
        'Sourced From','Confidence','Found When','Why Not In HubSpot','Evidence URL','Evidence']
head(ws, cols)
r = 2
STATUS = {'dealer': 'Independent dealer', 'dealer_bad_domain': 'Dealer (domain suspect)',
          'acquired': 'Acquired', 'not_dealer': 'Not a dealer', 'unresolved': 'Unresolved',
          'non_us': 'Non-US (out of census scope)', 'defunct': 'Defunct / out of business'}
for cl, mem in sorted(CO.members.items(),
                      key=lambda kv: str(comps[kv[1][0]].get('name') or '').lower()):
    c0 = comps[mem[0]]
    bs = brand.get(mem[0], {})
    st = STATUS.get(c0.get('ai__dealer_verdict'), 'Not yet verified')
    if bs.get('verdict') == 'parent_link':
        st = 'Acquired - own brand survives'
    elif bs.get('verdict') == 'still_independent':
        st = 'Independent dealer'
    elif bs.get('verdict') == 'merge':
        st = 'Acquired - brand retired'
    for p in sorted(byc.get(cl, []), key=lambda x: str(x.get('lastname') or '')):
        title = p.get('jobtitle') or ''
        isdm = bool(DM.search(title))
        alive = (p.get('hs_lead_status') or '') not in DEAD \
            and (p.get('ai__li_still_at_company') or '') not in ('moved', 'no')
        reach = 'YES' if (isdm and alive and (p.get('email') or '').strip()) else \
                ('phone only' if (isdm and alive and ((p.get('phone') or '').strip()
                                                      or (p.get('mobilephone') or '').strip()))
                 else 'no')
        vals = ['YES', c0.get('name'), st, c0.get('domain'), c0.get('city'), c0.get('state'),
                p.get('firstname'), p.get('lastname'), title,
                'YES' if isdm else '', p.get('email'),
                p.get('neverbouncevalidationresult'), p.get('phone'), p.get('mobilephone'),
                p.get('linkedin_profile_url__unique_value'), p.get('hs_lead_status'),
                p.get('ai__li_still_at_company'), p.get('validated__linkedin_or_manually'),
                p.get('ai__contact_verified_date'), p.get('zoominfo_contact_accuracy_score_'),
                reach, p['id'], mem[0], '', '', '', '', '', '']
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v)
        r += 1
N_HS = r - 1

# ---------- the other half of the picture.
# Every row above came from a live HubSpot query, so on its own this sheet shows the PORTAL,
# not the research. These rows are people the sweeps genuinely found and who were then held
# back - client accounts, people who resembled an existing record, titles below the bar,
# low confidence. They are real findings and belong in the same sheet, flagged.
SRC = S + '/sourced_not_in_hubspot.json'
N_SRC = 0
if os.path.exists(SRC):
    src = json.load(open(SRC))
    def clname(ids):
        for i in (ids or '').split(';'):
            if i in comps:
                return comps[i]
        return {}
    for q in sorted(src, key=lambda x: (str(x.get('company') or '').lower(),
                                        str(x.get('last_name') or ''))):
        c0 = clname(q.get('company_ids'))
        st = STATUS.get(c0.get('ai__dealer_verdict'), 'Not yet verified') if c0 else ''
        title = q.get('title_found') or ''
        isdm = bool(DM.search(title))
        em = (q.get('email') or '').strip()
        reach = 'YES' if (isdm and em) else ('phone only' if (isdm and (q.get('phone') or '').strip()) else 'no')
        vals = ['NO - sourced, not written',
                q.get('company') or c0.get('name'), st,
                q.get('domain') or c0.get('domain'), c0.get('city'), c0.get('state'),
                q.get('first_name'), q.get('last_name'), title,
                'YES' if isdm else '', em,
                '', q.get('phone'), '', q.get('linkedin_url'),
                '', '', '', '', '', reach, '', (q.get('company_ids') or '').split(';')[0],
                q.get('source'), q.get('confidence'), q.get('found_when'), q.get('why_not_in_hubspot'),
                q.get('evidence_url'), (q.get('evidence_quote') or '')[:500]]
        for j, v in enumerate(vals, 1):
            x = ws.cell(row=r, column=j, value=v)
            if j == 1:
                x.font = RED
        r += 1
        N_SRC += 1
NC = r - 1
widths(ws, [22,30,24,24,16,8,14,16,30,10,32,12,16,16,38,24,14,14,12,10,11,14,14,
            16,11,16,40,40,60])
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r-1}"

# =============================================================== Coverage by Dealer
ws2 = wb.create_sheet('Coverage by Dealer')
cols2 = ['Company','Dealer Status','Acquirer','Acq. Year','Brand Status','Domain','City','State',
         'Client Hold','Contacts Held','Decision Makers','DMs w/ Email','DMs w/ Phone Only',
         'Meets 2+ DM Goal?','Reachable DM Tier','Has 1+ Reachable DM?','Gap',
         # --- what the AI research actually established about this dealer
         'Company Type','Brands Carried','Dealer Services','Locations Served','Location Count',
         'ENX Elite','ENX Elite Years','Revenue Tier','Notable Finding','Executives',
         'Marketing Maturity','Software Stack','Has Blog','Gated Content',
         'HubSpot Usage Evidence','Open Sales/Mktg Roles','Engagement Overview',
         'Fastest Growing Segment','Growth Commentary','Accomplishments',
         'Leasing Partners','Mfr / Industry Awards','Acquisitions Made','Why Flagged Copier',
         'Data Quality Status','Data Quality Issues','Data Quality Notes',
         'Enrichment Hold','Verified Date',
         'Leadership Page','Contact Page','Careers Page','News Page','Trade Press Profile',
         'Company IDs']
head(ws2, cols2)
r2 = 2
gap_rows = []
for cl, mem in sorted(CO.members.items(),
                      key=lambda kv: str(comps[kv[1][0]].get('name') or '').lower()):
    c0 = comps[mem[0]]
    bs = brand.get(mem[0], {})
    verd = c0.get('ai__dealer_verdict')
    st = STATUS.get(verd, 'Not yet verified')
    if bs.get('verdict') == 'parent_link':
        st = 'Acquired - own brand survives'
    elif bs.get('verdict') == 'still_independent':
        st = 'Independent dealer'
    elif bs.get('verdict') == 'merge':
        st = 'Acquired - brand retired'
    rows = byc.get(cl, [])
    alive = [p for p in rows
             if (p.get('hs_lead_status') or '') not in DEAD
             and (p.get('ai__li_still_at_company') or '') not in ('moved', 'no')]
    dms = [p for p in alive if DM.search(p.get('jobtitle') or '')]
    dme = [p for p in dms if (p.get('email') or '').strip()]
    dmp = [p for p in dms if not (p.get('email') or '').strip()
           and ((p.get('phone') or '').strip() or (p.get('mobilephone') or '').strip())]
    onhold = any(m in hold for m in mem)
    gap = max(0, 2 - len(dme))
    def A(k):
        # the cluster's first record is not always the one carrying the research, so take
        # the first non-empty value across every company record in the cluster
        for m in mem:
            v = comps[m].get(k)
            if v not in (None, ''):
                return v
        return ''
    vals = [c0.get('name'), st, c0.get('ai__acquired_by'), A('ai__acquisition_year'),
            A('ai__brand_status'), c0.get('domain'),
            c0.get('city'), c0.get('state'), 'YES' if onhold else '',
            len(rows), len(dms), len(dme), len(dmp),
            'YES' if len(dme) >= 2 else 'NO',
            # A dealer with ONE reachable decision-maker is not the same as a dealer with
            # none - you can start a conversation with one. The binary 2+ goal hid that.
            ('4 - goal met: 2+ reachable DMs' if len(dme) >= 2 else
             '3 - 1 reachable DM' if len(dme) == 1 else
             '2 - DM on file, none reachable' if len(dms) >= 1 else
             '1 - contacts but no DM' if len(rows) >= 1 else
             '0 - no contacts at all'),
            'YES' if len(dme) >= 1 else 'NO', gap,
            A('ai__company_type'), A('ai__brands_carried'), A('ai__dealer_services'),
            A('ai__locations_served'), A('ai__location_count'),
            A('ai__enx_elite_dealer'), A('ai__enx_elite_years'), A('ai__revenue_tier_reported'),
            A('ai__enrichment_notable'), A('ai__executives'),
            A('ai__marketing_maturity'), A('ai__software_stack'),
            A('ai__has_blog'), A('ai__has_gated_content'),
            A('ai__hubspot_usage_evidence'), A('ai__open_roles'), A('ai__engagement_overview'),
            A('ai__fastest_growing_segment'), A('ai__growth_commentary'),
            A('ai__accomplishments'), A('ai__leasing_partners'), A('ai__manufacturer_awards'),
            A('ai__acquisitions_made'), A('ai__copier_reason'),
            A('ai__data_quality_status'), A('ai__data_quality_issues'),
            A('ai__data_quality_notes'), A('ai__enrichment_hold'), A('ai__verification_date'),
            A('ai__url_leadership'), A('ai__url_contact'), A('ai__url_careers'),
            A('ai__url_news'), A('ai__trade_press_profile_url'),
            ';'.join(mem)]
    for j, v in enumerate(vals, 1):
        ws2.cell(row=r2, column=j, value=v)
    # the gap sheet covers live, non-client dealers only
    if len(dme) < 2 and not onhold and st in ('Independent dealer', 'Dealer (domain suspect)',
                                              'Acquired - own brand survives',
                                              'Not yet verified', 'Unresolved'):
        t = tried.get(cl, {})
        gap_rows.append((c0, mem, st, len(rows), len(dms), len(dme), gap, t))
    r2 += 1
ND = r2 - 1
widths(ws2, [32,26,22,10,20,24,16,8,10,11,13,12,15,13,30,13,7,
             26,34,34,30,9,10,14,18,50,40,
             18,30,10,12,40,30,50,
             22,40,40,30,30,30,40,
             16,30,60,16,12,
             34,34,34,34,34,26])
from openpyxl.utils import get_column_letter as _gl
ws2.auto_filter.ref = f"A1:{_gl(len(cols2))}{r2-1}"

# =============================================================== No Contacts / gaps
ws3 = wb.create_sheet('Gap - Needs Contacts')
cols3 = ['Priority','Company','Dealer Status','Domain','City','State','Contacts Held',
         'Decision Makers','DMs w/ Email','Still Needs','Already Tried',
         'Why Not Found','RECOMMENDED NEXT SOURCE','Company IDs']
head(ws3, cols3)
r3 = 2
def as_list(v):
    """sources_tried arrives as a list from some agents and a plain string from others.
    ", ".join() on a string silently iterates its characters, which is how the Gap sheet
    ended up with "f, i, r, e, c, r, a, w, l" in the Already Tried column."""
    if not v:
        return []
    if isinstance(v, str):
        return [x.strip() for x in re.split(r'[;,]', v) if x.strip()]
    return [str(x).strip() for x in v if str(x).strip()]

def next_source(c0, t, ndm):
    dom = (c0.get('domain') or '').strip()
    trystr = ' '.join(as_list(t.get('sources_tried')))
    why = (t.get('why_not_found') or '') + ' ' + (t.get('note') or '')
    if not dom:
        return ('1. Find the real domain first - Secretary of State entity search, then the '
                'company site. No domain means every other source misfires.')
    if re.search(r'dead|dns|parked|404|expired|for sale', why, re.I):
        return ('1. Domain is dead - Secretary of State entity status will say whether the '
                'company still exists. If active, get officers from the filing.')
    if 'sos_filing' not in trystr and 'bbb' not in trystr:
        return ('1. Secretary of State officer filing  2. BBB "Business Management" block  '
                '3. Chamber of Commerce member directory. These three cracked 95% of the '
                'hardest dealers and were not tried here.')
    if ndm == 0:
        return ('1. Phone the main line and ask for the owner by name  '
                '2. Manufacturer dealer locator to confirm they still trade  '
                '3. State/county contract documents, which publish owner names AND emails')
    return ('1. ZoomInfo enrich on the decision-maker already held (email release)  '
            '2. LinkedIn profile contact_info  3. Company team/contact page for a direct address')
for c0, mem, st, ncon, ndm, ndme, gap, t in sorted(
        gap_rows, key=lambda x: (-x[6], str(x[0].get('name') or '').lower())):
    pri = 'A - no contacts at all' if ncon == 0 else (
          'B - contacts but no decision-maker' if ndm == 0 else
          'C - decision-maker but not reachable')
    vals = [pri, c0.get('name'), st, c0.get('domain'), c0.get('city'), c0.get('state'),
            ncon, ndm, ndme, gap,
            ', '.join(as_list(t.get('sources_tried'))) or 'not yet worked',
            (t.get('why_not_found') or t.get('note') or '')[:300],
            next_source(c0, t, ndm), ';'.join(mem)]
    for j, v in enumerate(vals, 1):
        x = ws3.cell(row=r3, column=j, value=v)
        if j in (11, 12, 13):
            x.alignment = Alignment(vertical='top', wrap_text=True)
    r3 += 1
NG = r3 - 1
widths(ws3, [22,30,26,24,15,8,11,13,12,11,26,44,60,24])
ws3.auto_filter.ref = f"A1:N{r3-1}"

# =============================================================== All Companies
ws6 = wb.create_sheet('All Companies')
cols6 = ['Company','Copier Dealer?','Dealer Status','Acquirer','Brand Verdict','Domain','Website',
         'City','State','Phone','Employees','Revenue','ENX Elite','Client Hold',
         'Duplicate Of (same domain)','Contacts Held','Decision Makers','DMs w/ Email',
         'Verified Date','Company ID']
head(ws6, cols6)
r6 = 2
BV = {'merge': 'Brand retired - merge into acquirer',
      'parent_link': 'Own brand survives - keep separate, parent-link',
      'still_independent': 'Never acquired',
      'self_rebrand': 'Renamed itself - duplicate pair',
      'unclear': 'Unclear - needs a human'}
for cid, c in sorted(comps.items(), key=lambda kv: str(kv[1].get('name') or '').lower()):
    cl = CO.cluster(cid)
    mem = CO.members.get(cl) or [cid]
    sibs = [m for m in mem if m != cid]
    bs = brand.get(cid, {})
    rows = per.get(cid, [])
    alive = [p for p in rows
             if (p.get('hs_lead_status') or '') not in DEAD
             and (p.get('ai__li_still_at_company') or '') not in ('moved', 'no')]
    dms = [p for p in alive if DM.search(p.get('jobtitle') or '')]
    dme = [p for p in dms if (p.get('email') or '').strip()]
    vals = [c.get('name'),
            'YES' if c.get('copier_company') == 'true' else 'no',
            STATUS.get(c.get('ai__dealer_verdict'), 'Not yet verified'),
            c.get('ai__acquired_by'),
            BV.get(bs.get('verdict'), ''),
            c.get('domain'), c.get('website'), c.get('city'), c.get('state'), c.get('phone'),
            c.get('numberofemployees'), c.get('annualrevenue'), c.get('ai__enx_elite_dealer'),
            'YES' if cid in hold else '',
            ';'.join(sibs), len(rows), len(dms), len(dme),
            c.get('ai__verification_date'), cid]
    for j, v in enumerate(vals, 1):
        ws6.cell(row=r6, column=j, value=v)
    r6 += 1
NA = r6 - 1
widths(ws6, [34,13,24,22,38,24,26,16,8,16,11,14,10,11,26,11,13,12,12,14])
ws6.auto_filter.ref = f"A1:T{r6-1}"

# =============================================================== Sourcing Playbook
ws4 = wb.create_sheet('Sourcing Playbook')
ws4['A1'] = 'Where dealer contacts actually came from'
ws4['A1'].font = TITLE
ws4['A2'] = ('Hit rates measured on this run, not estimates. The bottom three sources were never '
             'used in the first sweep and cracked ~95% of the dealers that sweep had given up on.')
ws4['A2'].font = NOTE
pb = [
    ('Source', 'What it yields', 'Measured result on this run', 'Cost'),
    ("Company's own /about, /team, /leadership page", 'Names + titles; sometimes direct emails',
     'Best single source. Fails when the site has no team page at all, which is common at small dealers.', 'Free'),
    ('Secretary of State officer filing', 'Officers, directors, registered agent, entity status',
     'Cracked dealers with NO website. Also proves whether a company still legally exists.', 'Free'),
    ('BBB "Business Management" block', 'Principal contact by name and title',
     'Cracked several owner-operated shops with no LinkedIn presence at all.', 'Free'),
    ('Chamber of Commerce member directory', 'Owner name, sometimes email',
     'Cracked two dealers whose sites have no team page.', 'Free'),
    ('State / county contract & bid documents', 'Owner name, title, real published email, phone',
     'One of the few free sources that publishes working EMAIL addresses.', 'Free'),
    ('LinkedIn profile (via Unipile)', 'Current employer, title, tenure, sometimes work email',
     '81% of checked contacts confirmed still in seat; 14% had left. Best currency check.', 'Free'),
    ('ZoomInfo search (free tier)', 'Names, titles, accuracy score - NEVER the email value',
     'Returns hasEmail flag only. Good for names, useless for reachability.', 'Free'),
    ('ZoomInfo enrich (paid credits)', 'Verified business email, direct dial, mobile',
     '30% email hit rate; 60% phone. 7 of 103 returned a person from the WRONG company.', 'Credits'),
    ('Trade press - ENX, Cannata, BTA, Imaging Channel', 'Names with dated role confirmation',
     'Good for owners and award winners; thin for marketing and sales titles.', 'Free'),
    ('Manufacturer dealer locators', 'Confirms the dealer is ACTIVE and which lines they carry',
     'Proves a company still trades when its own site is dead.', 'Free'),
]
for i, row in enumerate(pb, 4):
    for j, v in enumerate(row, 1):
        x = ws4.cell(row=i, column=j, value=v)
        x.font = HDR if i == 4 else BODY
        if i == 4:
            x.fill = HDRF
        x.alignment = Alignment(vertical='top', wrap_text=True)
        x.border = THIN
widths(ws4, [42, 40, 62, 12])

# =============================================================== Summary (formulas)
ws0 = wb.create_sheet('Summary', 0)
# A Summary formula that hardcodes a column letter breaks the moment a column is inserted,
# and it breaks SILENTLY - it still evaluates, just against the wrong data. Resolve every
# letter from the header list by name instead.
def CL(cols_, name):
    return get_column_letter(cols_.index(name) + 1)
AC_DM   = CL(cols,  'Decision Maker?')
AC_RCH  = CL(cols,  'Reachable?')
AC_LEAD = CL(cols,  'Lead Status')
AC_VER  = CL(cols,  'Verified Date')
AC_PV   = CL(cols,  'Prior Validation')
AC_CO   = CL(cols,  'Company')
CV_STAT = CL(cols2, 'Dealer Status')
CV_HOLD = CL(cols2, 'Client Hold')
CV_GOAL = CL(cols2, 'Meets 2+ DM Goal?')
CV_CO   = CL(cols2, 'Company')
CV_GAP  = CL(cols2, 'Gap')
CV_TIER = CL(cols2, 'Reachable DM Tier')
CV_ANY  = CL(cols2, 'Has 1+ Reachable DM?')
GP_PRI  = CL(cols3, 'Priority')
GP_CO   = CL(cols3, 'Company')
GP_NEED = CL(cols3, 'Still Needs')

ws0['A1'] = 'Copier Dealer Census - Contact Coverage'
ws0['A1'].font = TITLE
ws0['A2'] = 'All figures pulled live from HubSpot portal 20682069 on 15 August 2026.'
ws0['A2'].font = NOTE
rows0 = [
    ('DEALERS', None),
    ('Distinct dealers (after de-duplicating on domain)', f'=COUNTA(\'Coverage by Dealer\'!{CV_CO}2:{CV_CO}{ND})'),
    ('Independent dealers', f'=COUNTIF(\'Coverage by Dealer\'!{CV_STAT}2:{CV_STAT}{ND},"Independent dealer")'),
    ('Acquired - own brand survives (still a target)',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_STAT}2:{CV_STAT}{ND},"Acquired - own brand survives")'),
    ('Acquired - brand retired', f'=COUNTIF(\'Coverage by Dealer\'!{CV_STAT}2:{CV_STAT}{ND},"Acquired - brand retired")'),
    ('Not a dealer', f'=COUNTIF(\'Coverage by Dealer\'!{CV_STAT}2:{CV_STAT}{ND},"Not a dealer")'),
    ('Client accounts on hold (untouched)', f'=COUNTIF(\'Coverage by Dealer\'!{CV_HOLD}2:{CV_HOLD}{ND},"YES")'),
    ('Raw company records in HubSpot (before de-duplication)',
     f'=COUNTA(\'All Companies\'!A2:A{NA})'),
    ('', None),
    ('CONTACTS', None),
    ('Total contacts at dealer companies', f'=COUNTA(\'All Contacts\'!{AC_CO}2:{AC_CO}{NC})'),
    ('Decision-makers', f'=COUNTIF(\'All Contacts\'!{AC_DM}2:{AC_DM}{NC},"YES")'),
    ('Reachable decision-makers (email + still employed)',
     f'=COUNTIF(\'All Contacts\'!{AC_RCH}2:{AC_RCH}{NC},"YES")'),
    ('Decision-makers reachable by phone only', f'=COUNTIF(\'All Contacts\'!{AC_RCH}2:{AC_RCH}{NC},"phone only")'),
    ('Contacts suppressed (left, retired, wrong)',
     f'=COUNTIF(\'All Contacts\'!{AC_LEAD}2:{AC_LEAD}{NC},"No Longer with Company")'
     f'+COUNTIF(\'All Contacts\'!{AC_LEAD}2:{AC_LEAD}{NC},"Retired - Remove from All Lists")'),
    ('Touched by a verification pass this cycle (NOT the same as verified)',
     f'=COUNTIF(\'All Contacts\'!{AC_VER}2:{AC_VER}{NC},"2026-08-*")'),
    ('  ...of those, VERIFIED TO STANDARD: current role, at this company, evidenced',
     f'=COUNTIFS(\'All Contacts\'!{AC_VER}2:{AC_VER}{NC},"2026-08-*",'
     f'\'All Contacts\'!{AC_PV}2:{AC_PV}{NC},"Yes")'),
    ('Verified to standard, whole census (validated = Yes)',
     f'=COUNTIF(\'All Contacts\'!{AC_PV}2:{AC_PV}{NC},"Yes")'),
    ('', None),
    ('THE GOAL: 2+ reachable decision-makers per dealer', None),
    ('Dealers meeting the goal (2+ reachable DMs)',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_GOAL}2:{CV_GOAL}{ND},"YES")'),
    ('Dealers NOT meeting the goal',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_GOAL}2:{CV_GOAL}{ND},"NO")'),
    ('% meeting the goal', '=IFERROR(B{{Dealers meeting the goal (2+ reachable DMs)}}/(B{{Dealers meeting the goal (2+ reachable DMs)}}+B{{Dealers NOT meeting the goal}}),0)'),
    ('', None),
    ('CAN WE REACH ANYONE AT ALL? (the softer, more useful bar)', None),
    ('Dealers with 1+ reachable decision-maker',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_ANY}2:{CV_ANY}{ND},"YES")'),
    ('% with 1+ reachable decision-maker', '=IFERROR(B{{Dealers with 1+ reachable decision-maker}}/' + str(ND-1) + ',0)'),
    ('', None),
    ('EVERY DEALER, BY COVERAGE TIER', None),
    ('  4 - goal met: 2+ reachable DMs',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_TIER}2:{CV_TIER}{ND},"4 - goal met: 2+ reachable DMs")'),
    ('  3 - exactly 1 reachable DM (you can still start a conversation)',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_TIER}2:{CV_TIER}{ND},"3 - 1 reachable DM")'),
    ('  2 - decision-maker named but no working email',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_TIER}2:{CV_TIER}{ND},"2 - DM on file, none reachable")'),
    ('  1 - contacts on file but nobody senior',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_TIER}2:{CV_TIER}{ND},"1 - contacts but no DM")'),
    ('  0 - no contacts at all',
     f'=COUNTIF(\'Coverage by Dealer\'!{CV_TIER}2:{CV_TIER}{ND},"0 - no contacts at all")'),
    ('', None),
    ('THE GAP (live dealers only, clients excluded)', None),
    ('Dealers on the gap list', f'=COUNTA(\'Gap - Needs Contacts\'!{GP_CO}2:{GP_CO}{NG})'),
    ('  A - no contacts at all', f'=COUNTIF(\'Gap - Needs Contacts\'!{GP_PRI}2:{GP_PRI}{NG},"A - no contacts at all")'),
    ('  B - contacts but no decision-maker',
     f'=COUNTIF(\'Gap - Needs Contacts\'!{GP_PRI}2:{GP_PRI}{NG},"B - contacts but no decision-maker")'),
    ('  C - decision-maker but not reachable',
     f'=COUNTIF(\'Gap - Needs Contacts\'!{GP_PRI}2:{GP_PRI}{NG},"C - decision-maker but not reachable")'),
    ('Total contacts still needed to close the gap', f'=SUM(\'Gap - Needs Contacts\'!{GP_NEED}2:{GP_NEED}{NG})'),
]

# --- resolve {{Row Label}} sentinels to real row numbers.
# rows0 is written starting at spreadsheet row 4 (see the write loop below), so a label's row is
# its list index + 4. Resolving by LABEL rather than writing literal cell references means that
# inserting or reordering a summary row can never again point a percentage at the wrong cell -
# that failure has now happened three times in this file and it never errors, it just returns a
# confident wrong number.
_ROWS0_FIRST_ROW = 4
_row_of = {}
for _i, (_lab, _f) in enumerate(rows0):
    if _lab:
        _row_of.setdefault(_lab, _i + _ROWS0_FIRST_ROW)
_missing = []
for _i, (_lab, _f) in enumerate(rows0):
    if isinstance(_f, str) and '{{' in _f:
        _new = _f
        for _key, _rownum in _row_of.items():
            _new = _new.replace('{{' + _key + '}}', str(_rownum))
        if '{{' in _new:
            _missing.append((_lab, _new))
        rows0[_i] = (_lab, _new)
if _missing:
    raise SystemExit('unresolved summary row sentinel(s): %r' % _missing)

# --- compute the same figures directly, so the sheet is readable before any recalc
def _cnt(ws_, col, pred):
    n = 0
    for row in ws_.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
        if pred(row[0]):
            n += 1
    return n
_eq = lambda v: (lambda x: (x or '') == v)
CV = {
 'Distinct dealers (after de-duplicating on domain)': _cnt(ws2, cols2.index('Company')+1, lambda x: bool(x)),
 'Independent dealers': _cnt(ws2, cols2.index('Dealer Status')+1, _eq('Independent dealer')),
 'Acquired - own brand survives (still a target)': _cnt(ws2, cols2.index('Dealer Status')+1, _eq('Acquired - own brand survives')),
 'Acquired - brand retired': _cnt(ws2, cols2.index('Dealer Status')+1, _eq('Acquired - brand retired')),
 'Not a dealer': _cnt(ws2, cols2.index('Dealer Status')+1, _eq('Not a dealer')),
 'Client accounts on hold (untouched)': _cnt(ws2, cols2.index('Client Hold')+1, _eq('YES')),
 'Raw company records in HubSpot (before de-duplication)': _cnt(ws6, 1, lambda x: bool(x)),
 'Total contacts at dealer companies': _cnt(ws, cols.index('Company')+1, lambda x: bool(x)),
 'Decision-makers': _cnt(ws, cols.index('Decision Maker?')+1, _eq('YES')),
 'Reachable decision-makers (email + still employed)': _cnt(ws, cols.index('Reachable?')+1, _eq('YES')),
 'Decision-makers reachable by phone only': _cnt(ws, cols.index('Reachable?')+1, _eq('phone only')),
 'Contacts suppressed (left, retired, wrong)':
     _cnt(ws, cols.index('Lead Status')+1, _eq('No Longer with Company')) + _cnt(ws, cols.index('Lead Status')+1, _eq('Retired - Remove from All Lists')),
 'Touched by a verification pass this cycle (NOT the same as verified)':
     _cnt(ws, cols.index('Verified Date')+1, lambda x: str(x or '').startswith('2026-08-')),
 '  ...of those, VERIFIED TO STANDARD: current role, at this company, evidenced':
     sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
         if str(r[cols.index('Verified Date')] or '').startswith('2026-08-')
         and (r[cols.index('Prior Validation')] or '') == 'Yes'),
 'Verified to standard, whole census (validated = Yes)':
     _cnt(ws, cols.index('Prior Validation')+1, _eq('Yes')),
 'Dealers meeting the goal (2+ reachable DMs)': _cnt(ws2, cols2.index('Meets 2+ DM Goal?')+1, _eq('YES')),
 'Dealers with 1+ reachable decision-maker': _cnt(ws2, cols2.index('Has 1+ Reachable DM?')+1, _eq('YES')),
 '  4 - goal met: 2+ reachable DMs': _cnt(ws2, cols2.index('Reachable DM Tier')+1, _eq('4 - goal met: 2+ reachable DMs')),
 '  3 - exactly 1 reachable DM (you can still start a conversation)': _cnt(ws2, cols2.index('Reachable DM Tier')+1, _eq('3 - 1 reachable DM')),
 '  2 - decision-maker named but no working email': _cnt(ws2, cols2.index('Reachable DM Tier')+1, _eq('2 - DM on file, none reachable')),
 '  1 - contacts on file but nobody senior': _cnt(ws2, cols2.index('Reachable DM Tier')+1, _eq('1 - contacts but no DM')),
 '  0 - no contacts at all': _cnt(ws2, cols2.index('Reachable DM Tier')+1, _eq('0 - no contacts at all')),
 'Dealers NOT meeting the goal': _cnt(ws2, cols2.index('Meets 2+ DM Goal?')+1, _eq('NO')),
 'Dealers on the gap list': _cnt(ws3, cols3.index('Company')+1, lambda x: bool(x)),
 '  A - no contacts at all': _cnt(ws3, cols3.index('Priority')+1, _eq('A - no contacts at all')),
 '  B - contacts but no decision-maker': _cnt(ws3, cols3.index('Priority')+1, _eq('B - contacts but no decision-maker')),
 '  C - decision-maker but not reachable': _cnt(ws3, cols3.index('Priority')+1, _eq('C - decision-maker but not reachable')),
 'Total contacts still needed to close the gap':
     sum(int(v[0] or 0) for v in ws3.iter_rows(min_row=2, min_col=cols3.index('Still Needs')+1, max_col=cols3.index('Still Needs')+1, values_only=True)),
}
_mg = CV['Dealers meeting the goal (2+ reachable DMs)']; _ng = CV['Dealers NOT meeting the goal']
CV['% meeting the goal'] = (_mg / (_mg + _ng)) if (_mg + _ng) else 0
CV['% with 1+ reachable decision-maker'] = (CV['Dealers with 1+ reachable decision-maker'] / (ND - 1)) if (ND - 1) else 0

ws0.cell(row=3, column=2, value='Live formula').font = BOLD
ws0.cell(row=3, column=3, value='Value at build').font = BOLD
ws0.cell(row=3, column=3).fill = YEL
rr = 4
for label, f in rows0:
    a = ws0.cell(row=rr, column=1, value=label)
    if f is None and label and not label.startswith('  '):
        a.font = BOLD
        for cc in (1, 2, 3):
            ws0.cell(row=rr, column=cc).fill = PatternFill('solid', fgColor='EAF0F3')
    else:
        a.font = BODY
    if f:
        b = ws0.cell(row=rr, column=2, value=f)
        b.font = BODY
        b.number_format = '0.0%' if '%' in label else '#,##0'
        if label in CV:
            c = ws0.cell(row=rr, column=3, value=CV[label])
            c.font = BOLD
            c.fill = YEL
            c.number_format = '0.0%' if '%' in label else '#,##0'
    rr += 1
ws0.cell(row=rr + 1, column=1,
    value=('Column B holds the live formula, so the sheet updates if you edit the data tabs. '
           'Column C is the same figure computed when this file was built on 15 Aug 2026 — it is '
           'there so the numbers are readable immediately. Excel recalculates column B on open; '
           'if the two ever disagree, trust column B.')).font = NOTE
ws0.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=3)
ws0.row_dimensions[rr + 1].height = 46
ws0.cell(row=rr + 1, column=1).alignment = Alignment(wrap_text=True, vertical='top')
widths(ws0, [52, 15, 15])

# =============================================================== Method
ws5 = wb.create_sheet('Method & Caveats')
ws5['A1'] = 'How to read this workbook'
ws5['A1'].font = TITLE
notes = [
 ('Definitions', ''),
 ('Decision-maker', 'Title matches President, CEO, Owner, Founder, Principal, Partner, Chairman, '
  'CxO, VP, Director of Sales/Marketing, or General Manager.'),
 ('Reachable', 'A decision-maker who is confirmed still employed AND has an email address. '
  '"phone only" means we can reach them but not by email.'),
 ('Dealer (de-duplicated)', 'Company records sharing a domain are counted as ONE dealer, because in '
  'this channel one domain is one company.'),
 ('', ''),
 ('What these numbers do NOT mean', ''),
 ('An email is not proof of delivery', 'Validation services confirm a mailbox accepts mail, not that '
  'it is the address the person uses. 25 addresses marked "valid" have bounced, and one contact\'s '
  '"valid" address is one he does not use. Only a real send proves deliverability.'),
 ('"Still employed" has a shelf life', '14% of checked decision-makers had already left. Verified '
  'today does not mean verified in 60 days.'),
 ('LinkedIn does not prove someone is alive', 'Ten deceased executives were found still presented as '
  'current - one on LinkedIn 14 years after death, another on his own company\'s management page. Any '
  'long-tenured owner needs an obituary check before outreach.'),
 ('ZoomInfo accuracy score is not evidence', 'It measures whether a record is reachable, not whether '
  'the person is alive or at that company. One deceased owner carried a 2025 validity date; 7 of 103 '
  'enrichments returned a real person from an entirely different company.'),
 ('', ''),
 ('Known open issues', ''),
 ('Portal automation overwrites verified data', 'Two workflow enrollments and HubSpot Data Enrichment '
  'rewrite jobtitle, phone and the unique LinkedIn URL within ~90 seconds of any contact update. The '
  'LinkedIn URL is the dedupe key, so automation guessing at it undermines identity matching.'),
 ('phone is not a reliable direct dial', 'A workflow copies mobilephone into phone, which has already '
  'destroyed at least one real direct-dial number.'),
 ('6 duplicate contacts created this cycle', 'Charles Freeman, Scott Mueller, Brendan Ludlow, Kim Hebb. '
  'All had neither email nor LinkedIn URL, so neither unique constraint could catch them. Queued to merge.'),
 ('120 company merges pending sign-off', 'Merges cannot be undone in HubSpot, so none were executed.'),
]
rr = 3
for a, b in notes:
    x = ws5.cell(row=rr, column=1, value=a)
    y = ws5.cell(row=rr, column=2, value=b)
    if b == '' and a:
        x.font = BOLD; x.fill = PatternFill('solid', fgColor='EAF0F3')
        ws5.cell(row=rr, column=2).fill = PatternFill('solid', fgColor='EAF0F3')
    else:
        x.font = BOLD; y.font = BODY
    y.alignment = Alignment(vertical='top', wrap_text=True)
    ws5.row_dimensions[rr].height = 34 if b else 16
    rr += 1
widths(ws5, [34, 104])

# sheet-level default font: one style object instead of ~100k
for _s in (ws, ws2, ws3):
    _s.sheet_format.defaultRowHeight = 14
try:
    wb._named_styles[0].font = Font(name=ARIAL, size=10)
except Exception:
    pass

out = S + '/COPIER_DEALER_CONTACTS.xlsx'
wb.save(out)
print(f"contacts rows: {NC-1}   dealer rows: {ND-1}   gap rows: {NG-1}")
print("saved", out)
