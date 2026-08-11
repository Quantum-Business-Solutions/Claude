#!/usr/bin/env python3
"""Join scraped owner/founder data onto the master partner list.

Produces two deliverables:

  partner_master_with_owners.csv/.xlsx - all 7,444 directory partners, with
      owner columns filled in where one was found. Owner columns are blank for
      partners outside the Diamond/Platinum/Gold scope that was scraped, which
      is a coverage fact rather than a missing value.
  partner_owners_found.csv - just the partners where an owner was identified,
      for review before anything is loaded into HubSpot.

Names are filtered here rather than in the scraper so the raw output stays
auditable: the scraper's job is recall, this is where precision is applied.

Usage:
    python3 scripts/build_partner_owner_master.py
"""

from __future__ import annotations

import csv
import os
import re

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(REPO, "data")
PARTNERS = os.path.join(D, "hubspot_partners.csv")
SCRAPED = os.path.join(D, "partner_leadership_scraped.csv")
OUT_MASTER = os.path.join(D, "partner_master_with_owners.csv")
OUT_XLSX = os.path.join(D, "partner_master_with_owners.xlsx")
OUT_OWNERS = os.path.join(D, "partner_owners_found.csv")

# Words that appear in scraped strings which are page furniture, section
# headings or company names rather than parts of a person's name. A name
# containing any of these is dropped: "Meet Our Leader", "Team Mat",
# "Our Current Vacancies", "Business Benjamin".
NOT_PERSON = re.compile(
    r"\b(marketing|clients?|corporate|solutions?|revenue|operations?|agency|"
    r"digital|growth|inbound|sales|team|group|media|services?|consulting|"
    r"partners?|hubspot|contact|talk|view|read|more|our|your|let|lets|learn|"
    r"book|meet|story|stories|work|works|home|data|web|business|company|brand|"
    r"content|strategy|success|support|technology|software|vacancies|careers?|"
    r"leader|leaders|leadership|about|welcome|hello|discover|explore|current|"
    r"associate|expert|experts|specialist|manager|director|officer|founder|"
    r"owner|president|ceo|principal|chief)\b", re.I)

OWNER_COLS = ["owner_name", "owner_title", "owner_linkedin_url",
              "owner_confidence", "owner_source_page", "owner_extra_names",
              "site_emails"]


def is_person(name: str) -> bool:
    """Exactly two or three capitalised words, none of them page furniture."""
    n = (name or "").strip()
    words = n.split()
    if not 2 <= len(words) <= 3:
        return False
    if NOT_PERSON.search(n):
        return False
    # Every word must look like a name, not an ALLCAPS acronym or a stray digit.
    return all(re.fullmatch(r"[A-Z][A-Za-z'’À-ÿ-]{1,20}\.?", w) for w in words)


def load_owners() -> dict[str, list[dict]]:
    """domain -> owner rows, best first, junk names removed."""
    if not os.path.exists(SCRAPED):
        return {}
    by_domain: dict[str, list[dict]] = {}
    for r in csv.DictReader(open(SCRAPED, encoding="utf-8-sig")):
        if not is_person(r["name"]):
            continue
        by_domain.setdefault(r["domain"], []).append(r)
    rank = {"high": 0, "medium": 1, "": 2}
    kind = {"team": 0, "about": 1, "other": 2, "homepage": 3}
    for rows in by_domain.values():
        # A titled owner with a LinkedIn URL from a team page is the best row.
        rows.sort(key=lambda r: (rank.get(r["confidence"], 2),
                                 kind.get(r["page_kind"], 3),
                                 not r["linkedin_url"]))
    return by_domain


def main() -> None:
    owners = load_owners()
    partners = list(csv.DictReader(open(PARTNERS, encoding="utf-8-sig")))
    cols = list(partners[0].keys()) + OWNER_COLS

    rows, found_rows = [], []
    for p in partners:
        hits = owners.get(p["domain"], [])
        best = hits[0] if hits else None
        rec = dict(p)
        rec.update({
            "owner_name": best["name"] if best else "",
            "owner_title": best["titles"] if best else "",
            "owner_linkedin_url": best["linkedin_url"] if best else "",
            "owner_confidence": best["confidence"] if best else "",
            "owner_source_page": best["source_url"] if best else "",
            # Second and third names matter: co-founders are common, and an
            # extra name is a useful cross-check on the primary one.
            "owner_extra_names": "; ".join(
                f"{h['name']} ({h['titles']})" for h in hits[1:4]),
            "site_emails": best["site_emails"] if best else "",
        })
        rows.append(rec)
        if best:
            found_rows.append(rec)

    write_csv(OUT_MASTER, cols, rows)
    write_csv(OUT_OWNERS, cols, found_rows)
    write_xlsx(OUT_XLSX, cols, rows)

    tiers: dict[str, int] = {}
    for r in found_rows:
        tiers[r["tier"] or "untiered"] = tiers.get(r["tier"] or "untiered", 0) + 1
    print(f"partners in master list          {len(rows)}")
    print(f"partners with an owner identified {len(found_rows)}")
    print(f"  by tier                         " + ", ".join(
        f"{k}={v}" for k, v in sorted(tiers.items())))
    print(f"  with a LinkedIn URL             "
          f"{sum(1 for r in found_rows if r['owner_linkedin_url'])}")
    print(f"  with a site email               "
          f"{sum(1 for r in found_rows if r['site_emails'])}")
    print(f"  high confidence                 "
          f"{sum(1 for r in found_rows if r['owner_confidence'] == 'high')}")
    print(f"  second/third name also found    "
          f"{sum(1 for r in found_rows if r['owner_extra_names'])}")
    for path in (OUT_MASTER, OUT_OWNERS, OUT_XLSX):
        print(f"wrote {path} ({os.path.getsize(path) // 1024} KB)")


def write_csv(path: str, cols: list[str], rows: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_xlsx(path: str, cols: list[str], rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed - skipping xlsx")
        return
    # Control characters are legal in CSV but make openpyxl refuse to save.
    illegal = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([illegal.sub("", str(r.get(c, "") or "")) for c in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    for i, c in enumerate(cols, 1):
        width = max(len(c) + 2, 14)
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)
    wb.save(path)


if __name__ == "__main__":
    main()
