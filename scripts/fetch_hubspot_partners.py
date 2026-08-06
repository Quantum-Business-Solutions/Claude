#!/usr/bin/env python3
"""Build a master list of every partner in the HubSpot Solutions Partner Directory.

The public directory at https://ecosystem.hubspot.com/marketplace/solutions is a
client-rendered app backed by HubSpot's "chirp" RPC gateway. Two unauthenticated
public RPCs give us everything the directory shows:

  1. PersonalizationPublicRpc/search
       Paginated listing index. Sorted by LISTING_NAME so pagination is stable.
  2. MarketplaceListingDetailsRpc/getListingDetailsV3
       Full profile for one listing: website, offices, languages, services,
       industries, accreditations, certifications, company-size focus, budget.

Stage 2 is cached to disk as JSONL, so a interrupted run resumes instead of
re-fetching. Delete the cache to force a refresh.

Usage:
    python3 scripts/fetch_hubspot_partners.py
    python3 scripts/fetch_hubspot_partners.py --no-details   # index only, fast
    python3 scripts/fetch_hubspot_partners.py --limit 200    # smoke test
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import sys
import threading
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor

GATEWAY = "https://api.hubspot.com/chirp-frontend-external/v1/gateway"
SEARCH_RPC = (
    "com.hubspot.marketplace.personalization.rpc.PersonalizationPublicRpc/search"
)
DETAILS_RPC = (
    "com.hubspot.marketplace.listing.details.rpc."
    "MarketplaceListingDetailsRpc/getListingDetailsV3"
)

# Discriminant the gateway requires on filter-value unions.
STRING_FILTER = "com.hubspot.marketplace.search.models.filters.StringFilterQuery"

PROFILE_URL = "https://ecosystem.hubspot.com/marketplace/solutions/{slug}"

PAGE_SIZE = 250
DETAIL_WORKERS = 8
MAX_ATTEMPTS = 5

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_ROOT, "data")
CACHE_PATH = os.path.join(DATA_DIR, "listing_details.cache.jsonl")
CSV_PATH = os.path.join(DATA_DIR, "hubspot_partners.csv")
XLSX_PATH = os.path.join(DATA_DIR, "hubspot_partners.xlsx")


# --------------------------------------------------------------------------- #
# transport
# --------------------------------------------------------------------------- #

def call_rpc(path: str, payload: dict, timeout: int = 45) -> dict:
    """POST to the chirp gateway, retrying on transient failures."""
    body = json.dumps(payload).encode()
    last_err: Exception | None = None

    for attempt in range(MAX_ATTEMPTS):
        if attempt:
            # Exponential backoff with jitter so parallel workers desynchronise.
            time.sleep(min(2**attempt, 30) + random.random())
        req = urllib.request.Request(
            f"{GATEWAY}/{path}",
            data=body,
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Origin": "https://ecosystem.hubspot.com",
                "Referer": "https://ecosystem.hubspot.com/marketplace/solutions",
                "User-Agent": "qbs-partner-directory-export/1.0",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                parsed = json.loads(resp.read().decode())
        except urllib.error.HTTPError as exc:
            # 429/5xx are worth retrying; a 400 means our payload is wrong.
            if exc.code in (429, 500, 502, 503, 504):
                last_err = exc
                continue
            raise RuntimeError(
                f"{path} returned HTTP {exc.code}: {exc.read()[:400]!r}"
            ) from exc
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            last_err = exc
            continue

        if parsed.get("type") == "data":
            return parsed["data"]
        if parsed.get("type") == "rpcError":
            err = parsed["rpcError"]
            if err.get("shouldRetry"):
                last_err = RuntimeError(err.get("message", "retryable rpcError"))
                continue
            raise RuntimeError(f"{path} rpcError: {err.get('message')}")
        raise RuntimeError(f"{path} unexpected envelope: {json.dumps(parsed)[:400]}")

    raise RuntimeError(f"{path} failed after {MAX_ATTEMPTS} attempts: {last_err}")


def unwrap(node):
    """Flatten chirp's {"value": x, "__typename": ...} field wrappers."""
    if isinstance(node, dict):
        if "value" in node and "__typename" in node:
            return unwrap(node["value"])
        return {
            k: unwrap(v) for k, v in node.items() if k not in ("__typename", "@type")
        }
    if isinstance(node, list):
        return [unwrap(v) for v in node]
    return node


# --------------------------------------------------------------------------- #
# stage 1 — listing index
# --------------------------------------------------------------------------- #

def search_page(offset: int, length: int) -> dict:
    return call_rpc(
        SEARCH_RPC,
        {
            "filter": {
                "filterGroups": [
                    {
                        "filtersByField": {
                            "PRODUCT_TYPE": [
                                {
                                    "values": ["SOLUTIONS_PARTNER_PROFILE"],
                                    "clause": "OR",
                                    "negation": False,
                                    "__typename": STRING_FILTER,
                                }
                            ]
                        },
                        "clause": "AND",
                        "negation": False,
                    }
                ],
                "clause": "AND",
                "negation": False,
            },
            # Alphabetical sort keeps deep pagination stable; relevance ranking
            # would reshuffle between requests and drop/duplicate rows.
            "sorts": [{"field": "LISTING_NAME", "order": "ASC"}],
            "offset": offset,
            "length": length,
            "language": "en",
        },
    )


def fetch_index(limit: int | None) -> list[dict]:
    first = search_page(0, PAGE_SIZE)
    total = first["total"]
    target = min(total, limit) if limit else total
    print(f"directory reports {total} partner profiles; collecting {target}")

    by_id: dict[int, dict] = {}
    for card in first["cards"]:
        by_id[card["listingId"]] = card

    offset = PAGE_SIZE
    while len(by_id) < target and offset < total:
        page = search_page(offset, PAGE_SIZE)
        cards = page["cards"]
        if not cards:
            break
        for card in cards:
            by_id[card["listingId"]] = card
        offset += PAGE_SIZE
        print(f"  index: {len(by_id)}/{target}", end="\r", flush=True)

    print(f"  index: {len(by_id)} unique profiles collected")
    if not limit and len(by_id) != total:
        print(
            f"  ! warning: collected {len(by_id)} but directory reported {total}",
            file=sys.stderr,
        )

    rows = sorted(by_id.values(), key=lambda c: (c.get("listingName") or "").lower())
    return rows[:target] if limit else rows


# --------------------------------------------------------------------------- #
# stage 2 — per-profile detail (cached, resumable)
# --------------------------------------------------------------------------- #

def load_cache() -> dict[int, dict]:
    if not os.path.exists(CACHE_PATH):
        return {}
    cached: dict[int, dict] = {}
    with open(CACHE_PATH, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue  # tolerate a torn final line from a killed run
            if isinstance(rec.get("listingId"), int):
                cached[rec["listingId"]] = rec
    return cached


def fetch_details(listing_ids: list[int]) -> dict[int, dict]:
    cached = load_cache()
    todo = [i for i in listing_ids if i not in cached]
    print(f"details: {len(cached)} cached, {len(todo)} to fetch")
    if not todo:
        return cached

    os.makedirs(DATA_DIR, exist_ok=True)
    write_lock = threading.Lock()
    counter = {"done": 0, "failed": 0}

    with open(CACHE_PATH, "a", encoding="utf-8") as cache_fh:

        def worker(listing_id: int) -> None:
            try:
                data = call_rpc(DETAILS_RPC, {"listingId": listing_id, "language": "en"})
                rec = unwrap(data.get("listing")) or {}
                rec["listingId"] = listing_id
            except Exception as exc:  # noqa: BLE001 - keep going, report at end
                with write_lock:
                    counter["failed"] += 1
                    print(f"\n  ! {listing_id}: {exc}", file=sys.stderr)
                return
            with write_lock:
                cache_fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
                cache_fh.flush()
                cached[listing_id] = rec
                counter["done"] += 1
                if counter["done"] % 25 == 0:
                    print(
                        f"  details: {counter['done']}/{len(todo)}", end="\r", flush=True
                    )

        with ThreadPoolExecutor(max_workers=DETAIL_WORKERS) as pool:
            list(pool.map(worker, todo))

    print(f"  details: {counter['done']} fetched, {counter['failed']} failed")
    return cached


# --------------------------------------------------------------------------- #
# stage 3 — flatten
# --------------------------------------------------------------------------- #

COLUMNS = [
    "company_name",
    "domain",
    "website_is_platform_link",
    "directory_listing_name",
    "directory_url",
    "website",
    "tier",
    "partner_type",
    "review_count",
    "rating",
    "country",
    "city",
    "state",
    "all_countries",
    "regions",
    "office_model",
    "languages",
    "services",
    "industries",
    "accreditations",
    "certifications",
    "client_size_focus",
    "budget",
    "works_with_breeze",
    "integrations_supported",
    "first_published",
    "last_published",
    "listing_name",
    "slug",
    "listing_id",
    "partner_id",
    "profile_id",
    "partner_portal_id",
    "description",
]

TIER_RANK = {"elite": 0, "diamond": 1, "platinum": 2, "gold": 3}


def ms_to_date(value) -> str:
    if not isinstance(value, (int, float)) or value <= 0:
        return ""
    return time.strftime("%Y-%m-%d", time.gmtime(value / 1000))


sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from domainutil import (  # noqa: E402
    is_generic,
    needs_redirect_resolution,
    registrable_domain,
    resolve_final_domain,
)

# Populated by resolve_shortened_websites(): shortener URL -> real domain.
RESOLVED_DOMAINS: dict[str, str] = {}


def resolve_shortened_websites(details: dict[int, dict]) -> None:
    """Resolve partners whose website is a shortener into their real domain.

    Cached on disk because it costs one HTTP request per affected partner.
    """
    cache_path = os.path.join(DATA_DIR, "resolved_domains.json")
    if os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as fh:
            RESOLVED_DOMAINS.update(json.load(fh))

    todo = sorted({
        (d.get("companyUrl") or "")
        for d in details.values()
        if needs_redirect_resolution(d.get("companyUrl") or "")
    } - set(RESOLVED_DOMAINS))
    if not todo:
        return

    print(f"resolving {len(todo)} shortened partner websites")
    for i, url in enumerate(todo, 1):
        RESOLVED_DOMAINS[url] = resolve_final_domain(url)
        if i % 10 == 0:
            print(f"  resolved: {i}/{len(todo)}", end="\r", flush=True)

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(cache_path, "w", encoding="utf-8") as fh:
        json.dump(RESOLVED_DOMAINS, fh, indent=1, sort_keys=True)
    hit = sum(1 for v in RESOLVED_DOMAINS.values() if v)
    print(f"  resolved: {hit}/{len(RESOLVED_DOMAINS)} produced a real domain")

# Tokens that should not be title-cased when rebuilding a name from a slug.
SLUG_UPPER = {"llc", "inc", "ltd", "gmbh", "bv", "srl", "sa", "ag", "plc",
              "crm", "seo", "b2b", "b2c", "ai", "it", "uk", "usa", "us"}


def name_from_slug(slug: str) -> str:
    """Rebuild a plausible company name from the directory slug.

    The directory's companyName field is partner-editable marketing copy - some
    partners replace their name with a keyword list entirely (New Breed lists
    itself as "CRM Implementations, RevOps, AEO + Web, Demand Gen"). The slug is
    generated from the real company name when the profile is created, so it is a
    far better basis for a CRM-facing name.
    """
    if not slug:
        return ""
    words = [w for w in re.split(r"[-_]+", slug.strip().lower()) if w]
    out = []
    for w in words:
        out.append(w.upper() if w in SLUG_UPPER else w[:1].upper() + w[1:])
    return " ".join(out)


# Control characters that are legal in JSON but rejected by the XLSX writer.
# Partner-authored profile copy occasionally contains them.
ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def scrub(value):
    """Strip control characters and collapse whitespace in text values."""
    if not isinstance(value, str):
        return value
    return re.sub(r"\s+", " ", ILLEGAL_CHARS.sub("", value)).strip()


def join(values) -> str:
    if not values:
        return ""
    seen, out = set(), []
    for v in values:
        v = ("" if v is None else str(v)).strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return "; ".join(out)


def build_row(card: dict, detail: dict) -> dict:
    return {k: scrub(v) for k, v in _build_row(card, detail).items()}


def _build_row(card: dict, detail: dict) -> dict:
    products = card.get("products") or [{}]
    product = products[0] if products else {}
    reviews = card.get("reviewSummary") or {}

    # Profiles list offices under one of two keys depending on whether the
    # partner works from a physical office or remotely. Physical first, since
    # that is the closest thing the directory has to a headquarters.
    offices = [
        loc.get("physicalOfficeLocation") or {}
        for loc in (detail.get("physicalOfficeLocations") or [])
        if isinstance(loc, dict)
    ] + [
        loc.get("remoteLocation") or {}
        for loc in (detail.get("remoteLocations") or [])
        if isinstance(loc, dict)
    ]
    primary = offices[0] if offices else {}

    credentials = [c for c in (detail.get("credentials") or []) if isinstance(c, dict)]
    rating = reviews.get("overallRating")

    # Untiered profiles report the literal string "none".
    tier = (detail.get("tier") or product.get("partnerTier") or "").lower()
    if tier in ("none", "null"):
        tier = ""

    listed_name = detail.get("companyName") or card.get("companyName") or ""
    slug = card.get("slug") or detail.get("urlSlug") or ""
    website = detail.get("companyUrl") or ""

    return {
        # Slug-derived name is the CRM-facing one; the directory's own field is
        # kept alongside it because that is what HubSpot displays publicly.
        "company_name": name_from_slug(slug) or listed_name,
        # A shortener resolves to the real site; anything else uses the URL as
        # given. Partners who list only a LinkedIn/Linktree page have no usable
        # domain and cannot be domain-matched or imported.
        "domain": (
            RESOLVED_DOMAINS.get(website, "")
            if needs_redirect_resolution(website)
            else registrable_domain(website)
        ),
        "website_is_platform_link": is_generic(website),
        "directory_listing_name": listed_name,
        "directory_url": PROFILE_URL.format(slug=slug),
        "website": website,
        "tier": tier,
        "partner_type": detail.get("partnerType") or product.get("partnerType") or "",
        "review_count": reviews.get("reviewCount") or 0,
        "rating": round(rating, 2) if isinstance(rating, (int, float)) else "",
        "country": primary.get("country") or "",
        "city": primary.get("locality") or "",
        "state": primary.get("state") or "",
        "all_countries": join(o.get("country") for o in offices),
        "regions": join((detail.get("regionChoiceLabels") or {}).values()),
        "office_model": join((detail.get("officeLocationLabels") or {}).values()),
        "languages": join(detail.get("languages")),
        "services": join(
            s.get("name") for s in (detail.get("services") or []) if isinstance(s, dict)
        ),
        "industries": join((detail.get("industryChoiceLabels") or {}).values()),
        "accreditations": join(
            c.get("title") for c in credentials if c.get("type") == "ACCREDITATION"
        ),
        "certifications": join(
            c.get("title") for c in credentials if c.get("type") == "CERTIFICATION"
        ),
        "client_size_focus": join(
            (detail.get("companySizeSpecialtyLabels") or {}).values()
        ),
        "budget": join((detail.get("budgetChoiceLabels") or {}).values()),
        "works_with_breeze": bool(
            detail.get("worksWithBreeze") or product.get("worksWithBreeze")
        ),
        "integrations_supported": len(detail.get("integrations") or []),
        "first_published": ms_to_date(detail.get("firstPublishedAt")),
        "last_published": ms_to_date(detail.get("lastPublishedAt")),
        "listing_name": card.get("listingName") or "",
        "slug": card.get("slug") or "",
        "listing_id": card.get("listingId") or "",
        "partner_id": detail.get("sourceId") or product.get("sourceId") or "",
        "profile_id": detail.get("profileId") or "",
        "partner_portal_id": detail.get("partnerPortalId") or "",
        "description": (card.get("description") or "").replace("\n", " ").strip(),
    }


def sort_key(row: dict):
    """Tiered partners first (elite -> gold), then by review count."""
    return (
        TIER_RANK.get(row["tier"], 4),
        -(row["review_count"] or 0),
        row["company_name"].lower(),
    )


def write_csv(rows: list[dict]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"wrote {CSV_PATH} ({len(rows)} rows)")


def write_xlsx(rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed - skipping .xlsx (CSV still written)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "HubSpot Partners"
    ws.append([c.replace("_", " ").title() for c in COLUMNS])

    header_fill = PatternFill("solid", fgColor="FF7A59")  # HubSpot orange
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append([row.get(c, "") for c in COLUMNS])

    widths = {
        "company_name": 42,
        "directory_url": 52,
        "website": 40,
        "description": 80,
        "services": 50,
        "industries": 40,
        "certifications": 50,
        "accreditations": 34,
        "languages": 22,
        "regions": 30,
        "client_size_focus": 28,
    }
    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 16)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_PATH)
    print(f"wrote {XLSX_PATH} ({len(rows)} rows)")


def summarise(rows: list[dict]) -> None:
    print("\n--- summary ---")
    print(f"total partner profiles: {len(rows)}")

    tiers: dict[str, int] = {}
    for row in rows:
        tiers[row["tier"] or "(untiered)"] = tiers.get(row["tier"] or "(untiered)", 0) + 1
    print("by tier:")
    for tier in ["elite", "diamond", "platinum", "gold", "(untiered)"]:
        if tier in tiers:
            print(f"  {tier:12s} {tiers[tier]:>5}")

    countries: dict[str, int] = {}
    for row in rows:
        if row["country"]:
            countries[row["country"]] = countries.get(row["country"], 0) + 1
    top = sorted(countries.items(), key=lambda kv: -kv[1])[:10]
    print(f"countries represented: {len(countries)}")
    print("top 10: " + ", ".join(f"{c} ({n})" for c, n in top))

    with_site = sum(1 for r in rows if r["website"])
    reviewed = sum(1 for r in rows if r["review_count"])
    print(f"with website: {with_site}  |  with >=1 review: {reviewed}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--limit", type=int, help="only process the first N profiles")
    ap.add_argument(
        "--no-details",
        action="store_true",
        help="skip the per-profile enrichment pass",
    )
    args = ap.parse_args()

    cards = fetch_index(args.limit)
    listing_ids = [c["listingId"] for c in cards]

    details = {} if args.no_details else fetch_details(listing_ids)
    resolve_shortened_websites(details)

    rows = [build_row(card, details.get(card["listingId"], {})) for card in cards]
    rows.sort(key=sort_key)

    write_csv(rows)
    write_xlsx(rows)
    summarise(rows)
    return 0


if __name__ == "__main__":
    sys.exit(main())
